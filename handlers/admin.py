import logging
from io import BytesIO
from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, BufferedInputFile
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.filters import Command
import openpyxl
from openpyxl.styles import Font, PatternFill

import database as db
import keyboards as kb
from config import ADMIN_IDS

router = Router()
log = logging.getLogger(__name__)


def is_admin(uid: int) -> bool:
    return uid in ADMIN_IDS


# ── States ────────────────────────────────────────────────────────────────────

class Interview(StatesGroup):
    date     = State()
    time     = State()
    location = State()
    confirm  = State()


# ── /admin ────────────────────────────────────────────────────────────────────

@router.message(Command("admin"))
async def cmd_admin(message: Message, state: FSMContext):
    log.info(f"Admin urinish: {message.from_user.id} | ADMIN_IDS: {ADMIN_IDS}")
    if not is_admin(message.from_user.id):
        await message.answer(
            f"⛔ Sizning ID: <code>{message.from_user.id}</code>\n"
            f"Admin ro\'yxatida yo\'q.",
            parse_mode="HTML"
        )
        return
    await state.clear()
    stats = await db.get_stats()
    await message.answer(
        f"🛡️ <b>Admin panel</b>\n\n"
        f"📝 Jami arizalar: <b>{stats['total']}</b>\n"
        f"✅ Kelishini tasdiqlagan: <b>{stats['coming']}</b>\n"
        f"❌ Kela olmayman: <b>{stats['not_coming']}</b>\n"
        f"⏳ Javob bermagan: <b>{stats['no_reply']}</b>",
        parse_mode="HTML",
        reply_markup=kb.admin_menu_kb()
    )


# ── Arizalar ro'yxati ─────────────────────────────────────────────────────────

@router.message(F.text == "📋 Arizalar ro'yxati")
async def list_apps(message: Message):
    if not is_admin(message.from_user.id):
        return
    users = await db.get_all_applicants()
    if not users:
        await message.answer("Hozircha arizalar yo'q.")
        return

    text = f"📋 <b>Arizalar ({len(users)} ta):</b>\n\n"
    for i, u in enumerate(users, 1):
        reply_icon = {"yes": "✅", "no": "❌"}.get(u["interview_reply"], "⏳")
        text += (
            f"{i}. {reply_icon} <b>{u['full_name']}</b>\n"
            f"   👥 {u['guruh']} | 🎯 {u['interest']}\n"
            f"   📞 {u['phone']}\n\n"
        )
        if len(text) > 3800:
            text += "..."
            break

    await message.answer(text, parse_mode="HTML")


# ── Statistika ────────────────────────────────────────────────────────────────

@router.message(F.text == "📊 Statistika")
async def statistics(message: Message):
    if not is_admin(message.from_user.id):
        return
    stats = await db.get_stats()
    await message.answer(
        f"📊 <b>Statistika</b>\n\n"
        f"📝 Jami arizalar: <b>{stats['total']}</b>\n"
        f"✅ Kelishini tasdiqlagan: <b>{stats['coming']}</b>\n"
        f"❌ Kela olmayman degan: <b>{stats['not_coming']}</b>\n"
        f"⏳ Hali javob bermagan: <b>{stats['no_reply']}</b>",
        parse_mode="HTML"
    )


# ── Suhbat belgilash ──────────────────────────────────────────────────────────

@router.message(F.text == "📅 Suhbat belgilash")
async def suhbat_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    users = await db.get_all_applicants()
    if not users:
        await message.answer("Hozircha arizachilar yo'q.")
        return
    await state.set_state(Interview.date)
    await message.answer(
        f"👥 Jami <b>{len(users)}</b> ta arizachiga xabar yuboriladi.\n\n"
        f"📅 Suhbat sanasini kiriting:\n"
        f"<i>Misol: 03.05.2026</i>",
        parse_mode="HTML",
        reply_markup=kb.cancel_kb()
    )


@router.message(Interview.date)
async def suhbat_date(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Bekor qilindi.", reply_markup=kb.admin_menu_kb())
        return
    await state.update_data(date=message.text.strip())
    await state.set_state(Interview.time)
    await message.answer(
        "🕙 Suhbat vaqtini kiriting:\n<i>Misol: 10:00</i>",
        parse_mode="HTML"
    )


@router.message(Interview.time)
async def suhbat_time(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Bekor qilindi.", reply_markup=kb.admin_menu_kb())
        return
    await state.update_data(time=message.text.strip())
    await state.set_state(Interview.location)
    await message.answer(
        "📍 Suhbat joyini kiriting:\n<i>Misol: 3-qavat, 322-xona</i>",
        parse_mode="HTML"
    )


@router.message(Interview.location)
async def suhbat_location(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Bekor qilindi.", reply_markup=kb.admin_menu_kb())
        return
    await state.update_data(location=message.text.strip())
    data = await state.get_data()
    users = await db.get_all_applicants()
    await state.set_state(Interview.confirm)
    await message.answer(
        f"✅ <b>Tasdiqlash</b>\n\n"
        f"📅 Sana: <b>{data['date']}</b>\n"
        f"🕙 Soat: <b>{data['time']}</b>\n"
        f"📍 Joy: <b>{data['location']}</b>\n\n"
        f"👥 <b>{len(users)}</b> ta arizachiga yuboriladi\n\n"
        f"Davom etasizmi?",
        parse_mode="HTML",
        reply_markup=kb.interview_confirm_kb()
    )


@router.callback_query(F.data == "iv_send")
async def suhbat_send(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        return
    data = await state.get_data()
    users = await db.get_all_applicants()

    sent = 0
    failed = 0
    for u in users:
        first = u["full_name"].split()[0]
        try:
            await call.bot.send_message(
                u["telegram_id"],
                f"Assalomu alaykum, <b>{first}</b>! 👋\n\n"
                f"\"Intellectual Leaders Club\" suhbati haqida:\n\n"
                f"📅 Sana: <b>{data['date']}</b>\n"
                f"🕙 Soat: <b>{data['time']}</b>\n"
                f"📍 Joy: Termiz iqtisodiyot va servis universiteti,\n"
                f"<b>{data['location']}</b>\n\n"
                f"O'z vaqtida tashrif buyuring! ✨",
                parse_mode="HTML",
                reply_markup=kb.interview_reply_kb()
            )
            sent += 1
        except Exception as e:
            log.warning(f"Xabar yubormadi {u['telegram_id']}: {e}")
            failed += 1

    await db.save_interview(data["date"], data["time"], data["location"])
    result = f"✅ Xabar <b>{sent}</b> ta arizachiga yuborildi!"
    if failed:
        result += f"\n⚠️ {failed} ta foydalanuvchiga yuborilmadi."
    await call.message.answer(result, parse_mode="HTML", reply_markup=kb.admin_menu_kb())
    await state.clear()
    await call.answer()


@router.callback_query(F.data == "iv_cancel")
async def suhbat_cancel(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.message.answer("Bekor qilindi.", reply_markup=kb.admin_menu_kb())
    await call.answer()


# ── Excel eksport ─────────────────────────────────────────────────────────────

@router.message(F.text == "📤 Excel eksport")
async def excel_export(message: Message):
    if not is_admin(message.from_user.id):
        return
    users = await db.get_all_applicants()
    if not users:
        await message.answer("Hozircha arizalar yo'q.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arizalar"

    headers = [
        "#", "To'liq ism", "Fakultet", "Yo'nalish",
        "Guruh", "Telefon", "Qiziqish yo'nalishi",
        "Suhbat javobi", "Ariza sanasi"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5597")

    reply_map = {"yes": "✅ Keladi", "no": "❌ Kelmaidi", None: "⏳ Javob yo'q"}

    for i, u in enumerate(users, 1):
        ws.append([
            i,
            u["full_name"],
            u["fakultet"],
            u["yonalish"],
            u["guruh"],
            u["phone"],
            u["interest"],
            reply_map.get(u["interview_reply"]),
            str(u["created_at"])[:10]
        ])

    for col in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(length + 4, 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    await message.answer_document(
        BufferedInputFile(buf.read(), filename="ILC_Arizalar.xlsx"),
        caption=f"📊 Jami {len(users)} ta ariza"
    )
